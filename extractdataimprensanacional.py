import time
import re
import numpy as np
import openpyxl
from editais import BidBr
from datetime import datetime


from typing import List
from lista import get_list

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options


from selenium.common.exceptions import NoSuchElementException


# Retirar os duplicados
# Paginar
# Erro ao recuperar o numero do edital


SLEEP_TIME = 1
FLAG = False


def verifica_objeto(vobjeto, vobjeto_texto):
    if FLAG == True:
        return True

    Valor = True
    if vobjeto in {"equipamento", "equipamentos", "infraestrutura"}:
        conjunto_palavras = {
            " informatica",
            " informática",
            " ti ",
            " rede",
            " redes",
            " servidor ",
            " servidores ",
            " fio ",
            " software",
            " hardware",
        }
        numero_ocorrencias = sum(
            palavra in vobjeto_texto.lower() for palavra in conjunto_palavras
        )
        if numero_ocorrencias <= 0:
            return False

    if vobjeto in {"rede", "redes"}:
        conjunto_palavras = {
            "hospitalar",
            "hospital",
            "hospitalares",
            "elétrica",
            "elétricas",
            "eletrica",
            "eletricas",
            "municipal",
            "telefonia",
        }
        numero_ocorrencias = sum(
            palavra in vobjeto_texto.lower() for palavra in conjunto_palavras
        )
        if numero_ocorrencias > 0:
            return False

    return Valor


def encontrar_frase(frase_maior, frase_menor):
    frase_maior_lower = frase_maior.lower()
    frase_menor_lower = frase_menor.lower()
    return frase_menor_lower in frase_maior_lower


def extrair_dados(lista_objeto, data_inicio, data_fim) -> List[BidBr]:
    lista_total = []
    chrome_options = Options()
    # chrome_options.add_argument("--headless")

    driver = webdriver.Chrome(
        service=Service(),
        options=chrome_options,
    )
    lista_quantitativo = np.empty((0, 2))
    for objeto in lista_objeto:
        url_imprensa = f"https://www.in.gov.br/consulta/-/buscar/dou?q={objeto}&s=todos&exactDate=personalizado&sortType=0&delta=20&publishFrom={data_inicio.replace('/','-')}&publishTo={data_fim.replace('/','-')}"

        driver.get(url_imprensa)
        listaresultados = driver.find_elements(By.CLASS_NAME, "resultado")
        count_objeto = driver.find_element(
            By.XPATH, "//p[@class='search-total-label text-default']"
        ).text
        count = 0
        for item in listaresultados:
            (
                edital_num,
                orgao_nome,
                objeto_txt,
                website,
                uasg_num,
                link,
                data_abertura,
            ) = ("", "", "", "", "", "", "")

            if (
                encontrar_frase(item.text, "\nAVISO DE LICITAÇÃO") == True
                and verifica_objeto(objeto, item.text) == True
            ):
                try:
                    count += 1

                    lista_quantitativo = np.vstack(
                        (lista_quantitativo, np.array([objeto, count]))
                    )
                    texto_dado0 = item.find_element(By.CLASS_NAME, "title-marker")
                    texto_dado1 = texto_dado0.get_attribute("outerHTML")
                    href_regex = re.compile(r'href="([^"]+)"')
                    matches = href_regex.findall(texto_dado1)
                    link = matches[0]

                    website = "https://www.in.gov.br" + link

                    driver.execute_script("window.open('', '_blank');")
                    driver.switch_to.window(driver.window_handles[1])

                    driver.get(website)

                    # tem um erro aqui
                    edital_num = driver.find_element(
                        By.CSS_SELECTOR, "div[class='texto-dou'] p:nth-child(2)"
                    ).text

                    match = re.search(r"(?<=UASG )(\d+)(?![\d])", edital_num)
                    if match:
                        uasg_num = match[0]
                    else:
                        uasg_num = ""

                    orgao_nome = driver.find_element(
                        By.CLASS_NAME, "orgao-dou-data"
                    ).text
                    objeto_txt = driver.find_element(By.CLASS_NAME, "texto-dou").text

                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])

                except NoSuchElementException as e:
                    exception_message = str(e)
                    first_lines = "\n".join(exception_message.split("\n")[:3])
                    objeto_txt = f"\n --- Erro ao encontrar o aviso de licitacao ---- {first_lines}"
                lista_total.append(
                    BidBr(
                        orgao_nome,
                        edital_num,
                        uasg_num,
                        objeto_txt,
                        data_abertura,
                        objeto,
                        website,
                    )
                )

    totalizador = {}

    for row in lista_quantitativo:
        robjeto = row[0]
        rcount = row[1]

        if robjeto in totalizador:
            totalizador[robjeto] = max(totalizador[robjeto], rcount)
        else:
            totalizador[robjeto] = rcount

    for robjeto, rcount in totalizador.items():
        print(f"* {robjeto.upper()}    {rcount}")

    driver.close()
    return lista_total


def roda():
    inicio = time.time()
    todos_objeto = get_list()
    todos_objeto.sort()
    lista_novo = []

    data_inicio = datetime.now().strftime("%d/%m/%Y")
    data_fim = datetime.now().strftime("%d/%m/%Y")

    # data_inicio = "07/08/2023"
    # data_fim = "07/08/2023"

    print("Início\n")

    lista_novo = extrair_dados(todos_objeto, data_inicio, data_fim)
    count = len(lista_novo)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "MeuSheet"

    sheet.cell(row=1, column=1, value="ID")
    sheet.cell(row=1, column=2, value="Edital")
    sheet.cell(row=1, column=3, value="UASG")
    sheet.cell(row=1, column=4, value="Orgão")
    sheet.cell(row=1, column=5, value="Objeto Id")
    sheet.cell(row=1, column=6, value="Objeto")
    sheet.cell(row=1, column=7, value="Data Abertura")
    sheet.cell(row=1, column=8, value="Status")
    sheet.cell(row=1, column=9, value="WebSite")
    sheet.cell(row=1, column=10, value="OpenAI")

    count_id = 0
    idn = 0
    for obj in lista_novo:
        count_id += 1
        idn = count_id + 1

        sheet.cell(row=idn, column=1, value=count_id)
        sheet.cell(row=idn, column=2, value=obj.edital_num)
        sheet.cell(row=idn, column=3, value=obj.uasg_num)
        sheet.cell(row=idn, column=4, value=obj.orgao_nome)
        sheet.cell(row=idn, column=5, value=obj.objeto)
        sheet.cell(row=idn, column=6, value=obj.objeto_txt)
        sheet.cell(row=idn, column=7, value=obj.data_abertura)
        sheet.cell(row=idn, column=8, value="Não")
        sheet.cell(row=idn, column=9, value=obj.website).hyperlink = obj.website
        sheet.cell(row=idn, column=10, value="Nada")

    nome_arquivo = "arquivo_imprensa.xlsx"
    workbook.save(nome_arquivo)
    print(f"Arquivo '{nome_arquivo}' criado com sucesso!")

    print(f"\nTodas de Registro-> {count}")
    count = 0
    for obj in lista_novo:
        count += 1
        # print("\n")
        print("----------")
        print(f"{count}")
        print(f"Objeto: {obj.objeto.upper()}")
        print(f"Orgão: {obj.orgao_nome.upper()}")
        print(f"Edital: {obj.edital_num}")
        print(f"UASG: {obj.uasg_num}")
        print(f"Website: {obj.website}")
        print(f"Data: {obj.data_abertura}")
        print(f"Texto: {obj.objeto_txt[:530]}")

    print(f"\nTodas de Registro-> {count}")

    print("\nFim\n")
    fim = time.time()
    tempo_execucao = fim - inicio

    minutos = int(tempo_execucao // 60)
    secundos = int(tempo_execucao % 60)

    print(f"Tempo de execução: minuto {minutos} segundos {secundos}")


def main():
    roda()


if __name__ == "__main__":
    main()
