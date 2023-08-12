import time
import openpyxl
from typing import List
from editais import BidBr
from lista import get_list
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options

# nome da variáveis e funcoes uma merda
# paginar


# Constants
SLEEP_TIME = 1
FLAG = True


def extrair_dados(aobjeto, pdtaini, pdtafim) -> List[BidBr]:
    # paginar
    lista_editais = []
    lista_todos = []
    lista_ordenada = []

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver = webdriver.Firefox(options=chrome_options)

    for pobjeto in aobjeto:
        adados = List[BidBr]
        driver.get(
            "http://comprasnet.gov.br/ConsultaLicitacoes/ConsLicitacao_Filtro.asp"
        )
        driver.find_element(By.NAME, "txtObjeto").send_keys(pobjeto)

        driver.find_element(By.NAME, "dt_publ_ini").send_keys(pdtaini)
        driver.find_element(By.NAME, "dt_publ_fim").send_keys(pdtafim)

        driver.find_element(By.XPATH, "//input[@value='5']").click()
        driver.find_element(By.NAME, "chk_pregaoTodos").click()

        # problema com o send key
        driver.find_element(By.NAME, "ok").send_keys(Keys.ENTER)

        time.sleep(SLEEP_TIME)
        lista_editais = retira_dados(
            driver.find_elements(By.CLASS_NAME, "tex3"), pobjeto
        )
        for obj in lista_editais:
            lista_todos.append(
                BidBr(
                    obj.orgao_nome,
                    obj.edital_num,
                    obj.uasg_num,
                    obj.objeto_txt,
                    obj.data_abertura,
                    obj.objeto,
                    obj.website,
                )
            )

    lista_ordenada = retira_duplicado(lista_todos)

    print("\nTotal : ", len(lista_ordenada))
    print("\n")
    driver.close()
    return lista_ordenada


def retira_dados(lista_web_elements, pobjeto) -> List[BidBr]:
    lista_editais = []

    for element in lista_web_elements:
        orgao_nome, edital_num, uasg_num, objeto_txt, data_abertura = "", "", "", "", ""

        vdados = element.text.split("\n")

        for dado in vdados:
            if encontrar_palavra(dado, "UASG:"):
                uasg_num = dado[15:].strip()
            elif encontrar_palavra(dado, "Objeto:"):
                objeto_txt = dado
            elif encontrar_palavra(dado, "Pregão"):
                inicio = dado.find("º")
                fim = dado.find("-")
                edital_num = dado[inicio + 1 : fim].strip()
            elif encontrar_palavra(dado, "de:"):
                data_abertura = dado

            if encontrar_palavra(dado, "Histórico"):
                lista_editais.append(
                    BidBr(
                        orgao_nome,
                        edital_num,
                        uasg_num,
                        objeto_txt,
                        data_abertura,
                        pobjeto,
                        f"https://www.gov.br/compras/edital/{uasg_num}-5-{altera_numero_edital(edital_num)}",
                    )
                )

            if not uasg_num:
                orgao_nome += "\n" + dado

    if len(lista_editais) > 0:
        print("* " + pobjeto)

    return lista_editais


def retira_duplicado(lista_editais):
    lista_final = []

    old_uasg = ""
    old_edital = ""

    lista_editais_ordenada = sorted(
        lista_editais, key=lambda x: (x.EditalNum, x.UasgNum)
    )
    for obj in lista_editais_ordenada:
        if (not old_uasg == obj.uasg_num) and (not old_edital == obj.edital_num):
            if verifica_objeto(obj.objeto, obj.objeto_txt) == True:
                lista_final.append(
                    BidBr(
                        obj.orgao_nome,
                        obj.edital_num,
                        obj.uasg_num,
                        obj.objeto_txt,
                        obj.data_abertura,
                        obj.objeto,
                        f"https://www.gov.br/compras/edital/{obj.uasg_num}-5-{altera_numero_edital(obj.edital_num)}",
                    )
                )

            old_uasg = obj.uasg_num
            old_edital = obj.edital_num

    lista_editais_ordenada = sorted(lista_final, key=lambda x: (x.Objeto))

    return lista_editais_ordenada


# Funcoes
def encontrar_palavra(frase, palavra):
    frase_lower = frase.lower()
    palavra_lower = palavra.lower()
    palavras_na_frase = frase_lower.split()
    return palavra_lower in palavras_na_frase


def altera_numero_edital(codigo):
    texto = ""
    if codigo == "":
        return ""
    for i in range(10 - len(codigo)):
        texto += "0"
    return texto + codigo.replace("/", "-")


def verifica_objeto(vobjeto, vobjeto_texto):
    if FLAG == True:
        return True

    Valor = True
    if vobjeto in {"equipamento", "equipamentos", "infraestrutura"}:
        conjunto_palavras = {
            " informatica",
            " informática",
            " ti ",
            " rede",
            " redes",
            " servidor ",
            " servidores ",
            " fio ",
            " software",
            " hardware",
        }
        numero_ocorrencias = sum(
            palavra in vobjeto_texto.lower() for palavra in conjunto_palavras
        )
        if numero_ocorrencias <= 0:
            return False

    if vobjeto in {"rede", "redes"}:
        conjunto_palavras = {
            "hospitalar",
            "hospital",
            "hospitalares",
            "elétrica",
            "elétricas",
            "eletrica",
            "eletricas",
            "municipal",
            "telefonia",
        }
        numero_ocorrencias = sum(
            palavra in vobjeto_texto.lower() for palavra in conjunto_palavras
        )
        if numero_ocorrencias > 0:
            return False

    return Valor


def roda_rotina():
    inicio = time.time()
    str_search = get_list()

    data_inicio = datetime.now().strftime("%d/%m/%Y")
    data_fim = datetime.now().strftime("%d/%m/%Y")

    print("Início\n")
    # data_inicio = "25/07/2023"
    # data_fim = "25/07/2023"

    list_interna = extrair_dados(str_search, data_inicio, data_fim)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "MeuSheet"

    sheet.cell(row=1, column=1, value="ID")
    sheet.cell(row=1, column=2, value="Edital")
    sheet.cell(row=1, column=3, value="UASG")
    sheet.cell(row=1, column=4, value="Orgão")
    sheet.cell(row=1, column=5, value="Objeto Id")
    sheet.cell(row=1, column=6, value="Objeto")
    sheet.cell(row=1, column=7, value="Data Abertura")
    sheet.cell(row=1, column=8, value="Status")
    sheet.cell(row=1, column=9, value="WebSite")
    sheet.cell(row=1, column=10, value="OpenAI")

    count = 0
    count_objeto = 0
    old_objeto = ""
    cur_objeto = ""
    for obj in list_interna:
        count += 1
        idn = count + 1
        if old_objeto == "" or old_objeto == obj.objeto:
            count_objeto += 1
        elif old_objeto != obj.objeto:
            count_objeto = 1

        cur_objeto = obj.objeto
        obj.objeto = obj.objeto + " " + str(count_objeto)
        sheet.cell(row=idn, column=1, value=count)
        sheet.cell(row=idn, column=2, value=obj.edital_num)
        sheet.cell(row=idn, column=3, value=obj.uasg_num)
        sheet.cell(row=idn, column=4, value=obj.orgao_nome)
        sheet.cell(row=idn, column=5, value=obj.objeto)
        sheet.cell(row=idn, column=6, value=obj.objeto_txt)
        sheet.cell(row=idn, column=7, value=obj.data_abertura)
        sheet.cell(row=idn, column=8, value="Não")
        sheet.cell(row=idn, column=9, value=obj.website).hyperlink = obj.website
        sheet.cell(row=idn, column=10, value="Nada")
        old_objeto = cur_objeto

        print(f"{count}")
        print(f"{obj.get_value()}")

    nome_arquivo = "arquivo_conlicitacao.xlsx"
    workbook.save(nome_arquivo)
    print(f"Arquivo '{nome_arquivo}' criado com sucesso!")

    print("\nFim\n")
    fim = time.time()
    tempo_execucao = fim - inicio

    minutos = int(tempo_execucao // 60)
    segundos = int(tempo_execucao % 60)

    print(f"Tempo de execução: minuto {minutos} segundos {segundos}")


def main():
    roda_rotina()


if __name__ == "__main__":
    main()
